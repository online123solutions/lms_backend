from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .serializers import (
    TraineeSerializer, EmployeeSerializer, TrainerSerializer, AdminSerializer, LoginSerializer,UserExcelUploadSerializer,PasswordResetRequestSerializer,
    SOPSerializer,StandardLibraryItemSerializer, TraineeProfileUpdateSerializer, TrainerProfileUpdateSerializer, ChangePasswordSerializer
)
from .models import (
    CustomUser, TraineeProfile, EmployeeProfile, TrainerProfile, AdminProfile,NotificationReceipt,SOP,StandardLibraryItem
)
from .tasks import send_welcome_email,send_password_reset_email
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from rest_framework.authtoken.models import Token 
from django.contrib.auth import authenticate
from django.contrib.auth import login,logout
from rest_framework.permissions import IsAuthenticated,AllowAny
from rest_framework.authentication import TokenAuthentication
from rest_framework.authentication import TokenAuthentication
import openpyxl
from django.http import HttpResponse
from django.core.files.storage import default_storage
from rest_framework.parsers import MultiPartParser, FormParser
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from django.db import transaction
import logging
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_decode
from django.utils.encoding import force_str
from django.core.exceptions import PermissionDenied
from rest_framework.generics import ListAPIView
from django.db.models import Q
from .utils import get_user_department

logger = logging.getLogger(__name__)


class RegistrationView(APIView):
    @swagger_auto_schema(
        operation_description="Register a new user based on role (inactive until approved)",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["user"],
            properties={
                "user": openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "username": openapi.Schema(type=openapi.TYPE_STRING),
                        "email": openapi.Schema(type=openapi.TYPE_STRING, format='email'),
                        "password": openapi.Schema(type=openapi.TYPE_STRING),
                        "role": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            enum=["trainee", "employee", "trainer", "admin"]
                        ),
                    }
                ),
                "employee_id": openapi.Schema(type=openapi.TYPE_STRING),
                "department": openapi.Schema(type=openapi.TYPE_STRING),
                "designation": openapi.Schema(type=openapi.TYPE_STRING),
                "expertise": openapi.Schema(type=openapi.TYPE_STRING),  # For trainer
            }
        ),
        responses={
            201: openapi.Response("User registered successfully (inactive)"),
            400: "Bad Request",
            200: "User updated successfully (inactive)"
        },
    )
    def post(self, request, *args, **kwargs):
        user_data = request.data
        user_info = user_data.get('user', {})

        if not isinstance(user_info, dict):
            return Response({'error': 'Invalid user data format'}, status=status.HTTP_400_BAD_REQUEST)

        role = user_info.get('role')
        username = user_info.get('username')

        serializer_map = {
            'trainee': TraineeSerializer,
            'employee': EmployeeSerializer,
            'trainer': TrainerSerializer,
            'admin': AdminSerializer,
        }

        serializer_class = serializer_map.get(role)
        if not serializer_class:
            return Response({'error': 'Invalid role'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                user = CustomUser.objects.filter(username=username).first()
                if user:
                    user.email = user_info.get('email', user.email)
                    user.set_password(user_info.get('password'))
                    if user.is_active:
                        user.is_active = False
                    user.save()
                # Use serializer for create/update
                serializer = serializer_class(data=user_data, instance=user, context={'request': request})
                if serializer.is_valid():
                    user_profile = serializer.save()
                    logger.info(f"Profile saved for user {username}: {user_profile}")
                    send_welcome_email.delay(user_info.get('email'))
                    status_code = status.HTTP_201_CREATED if not user else status.HTTP_200_OK
                    return Response(
                        {"message": f"Registration successful for {role}. Awaiting approval."},
                        status=status_code,
                    )
                else:
                    logger.error(f"Validation errors: {serializer.errors}")
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Exception during registration: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        

class LoginView(APIView):
    permission_classes = [AllowAny]
    @swagger_auto_schema(
        request_body=LoginSerializer,
        responses={200: 'Token', 400: 'Bad Request'}
    )
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data['user']

            if not user.is_active:
                return Response(
                    {"error": "Account is not active. Please contact support."}, 
                    status=status.HTTP_403_FORBIDDEN
                )

            login(request, user)
            token, created = Token.objects.get_or_create(user=user)

            # Determine role-based dashboard URL using user.role
            role = getattr(user, "role", "").lower()

            if role== "admin":
                dashboard_url = f"/admin-dashboard/{user.username}"
            elif role == "trainee":
                dashboard_url = f"/trainee-dashboard/{user.username}"
            elif role == "trainer":
                dashboard_url = f"/trainer-dashboard/{user.username}"
            elif role == "employee":
                dashboard_url = f"/employee-dashboard/{user.username}"
            else:
                dashboard_url = f"/dashboard/{user.username}"  # fallback`

            return Response({
                "token": token.key,
                "username": user.username,
                "role": user.role,
                "is_superuser": user.is_superuser,
                "dashboard_url": dashboard_url,
            }, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UserLogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        logout(request)
        return Response({"detail": "Successfully logged out."}, status=status.HTTP_200_OK)
    
class DownloadUserTemplate(APIView):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Import Template"

        # Header
        ws.append([
            'role',   
            'username',
            'email',
            'password',
            'name',
            'employee_id',
            'department',
            'designation'
        ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="user_upload_template.xlsx"'
        wb.save(response)
        return response

    
class UploadUsersExcelView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    @swagger_auto_schema(
        request_body=UserExcelUploadSerializer,  
        responses={200: "File processed successfully", 400: "Bad Request"},
    )
    def post(self, request):
        if 'excel_file' not in request.FILES:
            return Response({"error": "Excel file is required."}, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith('.xlsx'):
            return Response({"error": "Invalid file type. Upload a .xlsx file."}, status=status.HTTP_400_BAD_REQUEST)

        file_path = default_storage.save('temp/' + excel_file.name, excel_file)
        wb = openpyxl.load_workbook(default_storage.open(file_path))
        sheet = wb.active

        success_count = 0
        error_count = 0
        errors = []

        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Column order: role, username, email, password, name, employee_id, department, designation
                role, username, email, password, name, employee_id, department, designation = row

                if not all([role, username, email, password, name, employee_id, department, designation]):
                    raise ValueError("Missing required fields.")

                if role not in ['trainee', 'employee', 'trainer', 'admin']:
                    raise ValueError(f"Invalid role: {role}")

                # Convert password to string (Excel may read numbers as integers)
                password = str(password)
                
                print(f"DEBUG: Processing user {username} with password: {password}")

                # Check if user exists by email or username (both are unique)
                user = CustomUser.objects.filter(Q(email=email) | Q(username=username)).first()
                
                if user:
                    # User already exists, skip this row
                    errors.append(f"Row {index}: User already exists (email: {email}, username: {username})")
                    error_count += 1
                    continue

                # Create new user with properly hashed password
                user = CustomUser.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    role=role,
                    is_active=False
                )
                
                print(f"DEBUG: Created user {username}, password hash starts with: {user.password[:20]}")

                # Convert string fields to strings (Excel may read them as other types)
                name = str(name) if name else ""
                employee_id = str(employee_id) if employee_id else ""
                department = str(department) if department else ""
                designation = str(designation) if designation else ""

                # Update or create profile (signal may have already created empty profile)
                if role == 'trainee':
                    TraineeProfile.objects.update_or_create(
                            user=user,
                        defaults={
                            'name': name,
                            'employee_id': employee_id,
                            'department': department,
                            'designation': designation
                        }
                        )
                elif role == 'employee':
                    EmployeeProfile.objects.update_or_create(
                            user=user,
                        defaults={
                            'name': name,
                            'employee_id': employee_id,
                            'department': department,
                            'designation': designation
                        }
                        )
                elif role == 'trainer':
                    TrainerProfile.objects.update_or_create(
                            user=user,
                        defaults={
                            'name': name,
                            'employee_id': employee_id,
                            'department': department,
                            'designation': designation
                        }
                        )
                elif role == 'admin':
                    AdminProfile.objects.update_or_create(
                            user=user,
                        defaults={
                            'name': name,
                            'employee_id': employee_id,
                            'department': department,
                            'designation': designation
                        }
                        )

                send_welcome_email.delay(email)
                success_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f"Row {index}: {str(e)}")

        return Response({
            "message": f"{success_count} user(s) registered successfully.",
            "errors": errors
        }, status=status.HTTP_200_OK)


class BulkPasswordResetView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    @swagger_auto_schema(
        request_body=UserExcelUploadSerializer,  
        responses={200: "Passwords updated successfully", 400: "Bad Request"},
    )
    def post(self, request):
        if 'excel_file' not in request.FILES:
            return Response({"error": "Excel file is required."}, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith('.xlsx'):
            return Response({"error": "Invalid file type. Upload a .xlsx file."}, status=status.HTTP_400_BAD_REQUEST)

        file_path = default_storage.save('temp/' + excel_file.name, excel_file)
        wb = openpyxl.load_workbook(default_storage.open(file_path))
        sheet = wb.active

        updated_count = 0
        not_found_count = 0
        errors = []

        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Column order: role, username, email, password, name, employee_id, department, designation
                role, username, email, password, name, employee_id, department, designation = row

                if not username or not password:
                    continue

                # Convert password to string
                password = str(password)

                # Find user by username
                try:
                    user = CustomUser.objects.get(username=username)
                    # Set the password properly (this will hash it)
                    user.set_password(password)
                    user.save()
                    updated_count += 1
                except CustomUser.DoesNotExist:
                    not_found_count += 1
                    errors.append(f"Row {index}: User not found - {username}")

            except Exception as e:
                errors.append(f"Row {index}: {str(e)}")

        return Response({
            "message": f"{updated_count} password(s) updated successfully.",
            "not_found": not_found_count,
            "errors": errors
        }, status=status.HTTP_200_OK)

    
@swagger_auto_schema(
    method='post',
    operation_description="Mark a notification as read for the current user.",
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={"notification_id": openapi.Schema(type=openapi.TYPE_INTEGER)},
        required=["notification_id"]
    ),
    responses={200: "OK", 404: "Not Found"}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_notification_read(request):
    nid = request.data.get('notification_id')
    if not nid:
        return Response({"error": "notification_id is required"}, status=400)
    try:
        rec = NotificationReceipt.objects.get(notification_id=nid, user=request.user)
    except NotificationReceipt.DoesNotExist:
        return Response({"error": "Notification not found for this user."}, status=404)

    if not rec.is_read:
        rec.is_read = True
        rec.read_at = timezone.now()
        rec.save(update_fields=['is_read', 'read_at'])

    return Response({"message": "Marked as read."}, status=200)

class PasswordResetRequestView(APIView):
    permission_classes = [AllowAny]
    
    def initial(self, request, *args, **kwargs):
        self.raw_body = request.body.decode('utf-8') if request.body else 'No body'
        return super().initial(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Request a password reset email.",
        request_body=PasswordResetRequestSerializer,
        responses={200: "Password reset email is being sent", 400: "Bad Request"}
    )
    def post(self, request):

        try:
            serializer = PasswordResetRequestSerializer(data=request.data)
            if serializer.is_valid():
                email = serializer.validated_data.get('email')
                if not email:
                    return Response({"error": "Validation failed to provide an email."}, status=status.HTTP_400_BAD_REQUEST)
                send_password_reset_email.delay(email)
                return Response({'success': 'Password reset email is being sent'}, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"detail": "An unexpected error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
class PasswordResetConfirmView(APIView):
    permission_classes = [AllowAny]
    
    def post(self, request, uidb64, token):
        try:
            uid = force_str(urlsafe_base64_decode(uidb64))
            user = CustomUser.objects.get(pk=uid)
            if not default_token_generator.check_token(user, token):
                return Response({"detail": "Invalid or expired token."}, status=status.HTTP_400_BAD_REQUEST)

            new_password = request.data.get('new_password')
            if not new_password:
                return Response({"detail": "New password is required."}, status=status.HTTP_400_BAD_REQUEST)
            user.set_password(new_password)
            user.save()
            return Response({"detail": "Password has been reset successfully."}, status=status.HTTP_200_OK)
        except (TypeError, ValueError, OverflowError) as e:
            return Response({"detail": "Invalid token or user ID."}, status=status.HTTP_400_BAD_REQUEST)
        except CustomUser.DoesNotExist as e:
            return Response({"detail": "Invalid token or user ID."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"detail": "An unexpected error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class ChangePasswordView(APIView):
    """
    API endpoint for any authenticated user (trainee, trainer, employee, admin) 
    to change their own password.
    Requires old password and new password confirmation.
    No email required - direct password change in the app.
    """
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_description="Change password for the authenticated user. Requires old password and new password confirmation. Available for all roles.",
        request_body=ChangePasswordSerializer,
        responses={
            200: "Password changed successfully",
            400: "Bad Request - Invalid old password or passwords don't match",
            401: "Unauthorized"
        },
    )
    def post(self, request):
        """Change password for the authenticated user"""
        serializer = ChangePasswordSerializer(data=request.data, context={'request': request})
        
        if serializer.is_valid():
            user = request.user
            new_password = serializer.validated_data['new_password']
            
            # Set new password
            user.set_password(new_password)
            user.save()
            
            return Response(
                {"detail": "Password has been changed successfully."},
                status=status.HTTP_200_OK
            )
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class TraineeProfileUpdateView(APIView):
    """
    API endpoint for trainees to update their own profile.
    Only allows updating: name, department, designation, and profile_picture
    Supports both Token and Session authentication.
    """
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    @swagger_auto_schema(
        operation_description="Get current trainee profile. Only the authenticated trainee can access their own profile.",
        responses={
            200: openapi.Response("Profile retrieved successfully", TraineeProfileUpdateSerializer),
            403: "Forbidden - Only trainees can access this endpoint",
            404: "Profile not found"
        },
    )
    def get(self, request):
        """Get current trainee profile"""
        if request.user.role != 'trainee':
            return Response(
                {"detail": "Only trainees can access this endpoint."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            profile = TraineeProfile.objects.get(user=request.user)
            serializer = TraineeProfileUpdateSerializer(profile, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        except TraineeProfile.DoesNotExist:
            return Response(
                {"detail": "Profile not found."},
                status=status.HTTP_404_NOT_FOUND
            )

    @swagger_auto_schema(
        operation_description="Update trainee profile. Only the authenticated trainee can update their own profile.",
        request_body=TraineeProfileUpdateSerializer,
        responses={
            200: openapi.Response("Profile updated successfully", TraineeProfileUpdateSerializer),
            400: "Bad Request",
            403: "Forbidden - Only trainees can access this endpoint",
            404: "Profile not found"
        },
    )
    def put(self, request):
        """Update trainee profile (full update)"""
        if request.user.role != 'trainee':
            return Response(
                {"detail": "Only trainees can access this endpoint."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            profile = TraineeProfile.objects.get(user=request.user)
            serializer = TraineeProfileUpdateSerializer(
                profile, 
                data=request.data, 
                partial=False,
                context={'request': request}
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TraineeProfile.DoesNotExist:
            return Response(
                {"detail": "Profile not found."},
                status=status.HTTP_404_NOT_FOUND
            )

    @swagger_auto_schema(
        operation_description="Partially update trainee profile. Only the authenticated trainee can update their own profile.",
        request_body=TraineeProfileUpdateSerializer,
        responses={
            200: openapi.Response("Profile updated successfully", TraineeProfileUpdateSerializer),
            400: "Bad Request",
            403: "Forbidden - Only trainees can access this endpoint",
            404: "Profile not found"
        },
    )
    def patch(self, request):
        """Partially update trainee profile"""
        if request.user.role != 'trainee':
            return Response(
                {"detail": "Only trainees can access this endpoint."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            profile = TraineeProfile.objects.get(user=request.user)
            serializer = TraineeProfileUpdateSerializer(
                profile, 
                data=request.data, 
                partial=True,
                context={'request': request}
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TraineeProfile.DoesNotExist:
            return Response(
                {"detail": "Profile not found."},
                status=status.HTTP_404_NOT_FOUND
            )


class TrainerProfileUpdateView(APIView):
    """
    API endpoint for trainers to update their own profile.
    Only allows updating: name, department, designation, expertise, and profile_picture
    Supports both Token and Session authentication.
    """
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    @swagger_auto_schema(
        operation_description="Get trainer profile. Only the authenticated trainer can access their own profile.",
        responses={
            200: openapi.Response("Profile retrieved successfully", TrainerProfileUpdateSerializer),
            403: "Forbidden - Only trainers can access this endpoint",
            404: "Profile not found"
        },
    )
    def get(self, request):
        """Get current trainer profile"""
        if request.user.role != 'trainer':
            return Response(
                {"detail": "Only trainers can access this endpoint."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            profile = TrainerProfile.objects.get(user=request.user)
            serializer = TrainerProfileUpdateSerializer(profile, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        except TrainerProfile.DoesNotExist:
            return Response(
                {"detail": "Profile not found."},
                status=status.HTTP_404_NOT_FOUND
            )

    @swagger_auto_schema(
        operation_description="Update trainer profile. Only the authenticated trainer can update their own profile.",
        request_body=TrainerProfileUpdateSerializer,
        responses={
            200: openapi.Response("Profile updated successfully", TrainerProfileUpdateSerializer),
            400: "Bad Request",
            403: "Forbidden - Only trainers can access this endpoint",
            404: "Profile not found"
        },
    )
    def put(self, request):
        """Update trainer profile (full update)"""
        if request.user.role != 'trainer':
            return Response(
                {"detail": "Only trainers can access this endpoint."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            profile = TrainerProfile.objects.get(user=request.user)
            serializer = TrainerProfileUpdateSerializer(
                profile, 
                data=request.data, 
                partial=False,
                context={'request': request}
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TrainerProfile.DoesNotExist:
            return Response(
                {"detail": "Profile not found."},
                status=status.HTTP_404_NOT_FOUND
            )

    @swagger_auto_schema(
        operation_description="Partially update trainer profile. Only the authenticated trainer can update their own profile.",
        request_body=TrainerProfileUpdateSerializer,
        responses={
            200: openapi.Response("Profile updated successfully", TrainerProfileUpdateSerializer),
            400: "Bad Request",
            403: "Forbidden - Only trainers can access this endpoint",
            404: "Profile not found"
        },
    )
    def patch(self, request):
        """Partially update trainer profile"""
        if request.user.role != 'trainer':
            return Response(
                {"detail": "Only trainers can access this endpoint."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            profile = TrainerProfile.objects.get(user=request.user)
            serializer = TrainerProfileUpdateSerializer(
                profile, 
                data=request.data, 
                partial=True,
                context={'request': request}
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TrainerProfile.DoesNotExist:
            return Response(
                {"detail": "Profile not found."},
                status=status.HTTP_404_NOT_FOUND
            )


LENIENT_IF_NO_DEPT = True  # set False to be strict

def user_role(user):
    return (getattr(user, "role", "") or "").lower()

class BaseSOPListView(ListAPIView):
    serializer_class = SOPSerializer
    permission_classes = [IsAuthenticated]
    REQUIRED_ROLE = None  # 'trainer' | 'trainee' | 'employee' | 'admin'

    def has_required_role(self, user):
        return user.is_superuser or self.REQUIRED_ROLE is None or user_role(user) == self.REQUIRED_ROLE

    def get_queryset(self):
        u = self.request.user
        if not self.has_required_role(u):
            raise PermissionDenied("Not allowed for this role.")

        role = user_role(u)
        dept = get_user_department(u)  # '' if not found

        # INDIVIDUAL
        q_individual = Q(recipient_type="INDIVIDUAL", recipients=u)

        # GROUP
        if dept:
            dept_part = Q(department="") | Q(department__isnull=True) | Q(department__iexact=dept)
        else:
            # No department available for this user
            dept_part = Q() if LENIENT_IF_NO_DEPT else (Q(department="") | Q(department__isnull=True))

        # target_role in SOP is lower/upper tolerant
        if role:
            role_part = Q(target_role="") | Q(target_role__isnull=True) | Q(target_role__iexact=role)
        else:
            role_part = Q(target_role="") | Q(target_role__isnull=True)

        q_group = Q(recipient_type="GROUP") & dept_part & role_part

        return (
            SOP.objects.filter(q_individual | q_group)
            .select_related("created_by")
            .prefetch_related("recipients")
            .order_by("-created_at")
            .distinct()
        )

class BaseSLListView(ListAPIView):
    """
    Standard Library (group/global) list view.
    Filters by user's role/department:
      role_part = (target_role blank|NULL|iexact user_role)
      dept_part = (department blank|NULL|iexact user_dept)
      If no department on user and LENIENT_IF_NO_DEPT=True -> allow all departments.
    Supports ?q= (title/note) and ?tags=tag1,tag2 (JSONField overlap on PG, icontains fallback).
    """
    serializer_class = StandardLibraryItemSerializer
    permission_classes = [IsAuthenticated]
    REQUIRED_ROLE = None  # 'trainer' | 'trainee' | 'employee' | 'admin'

    def has_required_role(self, user):
        return user.is_superuser or self.REQUIRED_ROLE is None or user_role(user) == self.REQUIRED_ROLE

    def get_queryset(self):
        u = self.request.user
        if not self.has_required_role(u):
            raise PermissionDenied("Not allowed for this role.")

        role = (user_role(u) or "").strip()
        dept = (get_user_department(u) or "").strip()

        # Department filtering
        if dept:
            dept_part = Q(department="") | Q(department__isnull=True) | Q(department__iexact=dept)
        else:
            # If user has no department, either allow all departments (lenient),
            # or restrict to global (blank/NULL) only.
            dept_part = Q() if LENIENT_IF_NO_DEPT else (Q(department="") | Q(department__isnull=True))

        # Role filtering (case-insensitive, blanks allowed)
        if role:
            role_part = Q(target_role="") | Q(target_role__isnull=True) | Q(target_role__iexact=role)
        else:
            role_part = Q(target_role="") | Q(target_role__isnull=True)

        base_q = dept_part & role_part

        qs = (
            StandardLibraryItem.objects.filter(is_active=True)
            .filter(base_q)
            .select_related("created_by")
            .order_by("-created_at")
            .distinct()
        )

        # Optional text search
        q = (self.request.query_params.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(note__icontains=q))

        # Optional tags filter (?tags=a,b)
        tags_param = (self.request.query_params.get("tags") or "").strip()
        if tags_param:
            tags = [t.strip() for t in tags_param.split(",") if t.strip()]
            if tags:
                try:
                    qs = qs.filter(tags__overlap=tags)  # PostgreSQL JSONField
                except Exception:
                    for t in tags:  # SQLite/MySQL fallback
                        qs = qs.filter(tags__icontains=t)

        return qs
